"""Excel template, layout detection, normalization, and row parsing."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from pathlib import Path
from typing import cast

import openpyxl
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from Backend.app.domain.students import validate_nim_value
from Backend.db import parse_entry_registration
from Backend.excel_reader import (
    clean_demographic_value,
    clean_excel_text,
    normalize_imported_name,
    normalize_name,
    normalize_nim,
    normalize_text,
    read_sheet,
    read_sheet_headers,
    workbook_sheet_names,
)

DEFAULT_WORKBOOK = Path(__file__).resolve().parents[2] / "MASTER_DATA_2023_1_2026_1.xlsx"
DEFAULT_PERIOD = "UKT 2023.1 s/d 2025.2"
DEFAULT_BILL_TYPE = "UKT BRIVA"
DEFAULT_CURRENT_PERIOD = "Semester Ganjil 2026"
DEFAULT_INSTRUCTIONS = (
    "Bayar melalui BRIVA BRI dengan nomor VA yang tampil. "
    "Pastikan nama dan nominal sesuai sebelum menyelesaikan pembayaran."
)
REQUIRED_SYNC_HEADERS = ("NIM", "Nama Mahasiswa", "BRIVA", "Jumlah")
OPTIONAL_ISSUE_HEADERS = ("NIM", "Nama Mahasiswa", "BRIVA", "Jumlah", "Keterangan")
CURRENT_REQUIRED_HEADERS = ("NIM", "Nama", "No Rek", "Jumlah")

MASTER_TEMPLATE_HEADERS = [
    "NIM",
    "Nama",
    "NO KTP",
    "Tempat Lahir",
    "Tanggal Lahir",
    "Nama Ibu Kandung",
    "e-Mail",
    "No Kontak",
    "Registrasi Awal",
    "Program Studi",
    "No Rek",
    "Jumlah",
    "Batas Pembayaran",
]

MASTER_SAMPLE_ROWS = [
    [
        "000000001",
        "Mahasiswa Contoh Satu",
        "0000000000000001",
        "Kota Contoh",
        "01 Januari 2000",
        "Orang Tua Contoh",
        "mahasiswa.satu@example.test",
        "000000000001",
        "UNIVERSITAS TERBUKA 2023.1",
        "FEB - Akuntansi",
        "000000000000001",
        1850000,
        "22 Januari 2027 Pukul 11.59 WIB",
    ],
    [
        "000000002",
        "Mahasiswa Contoh Dua",
        "0000000000000002",
        "Kota Contoh",
        "02 Februari 2001",
        "Orang Tua Contoh",
        "mahasiswa.dua@example.test",
        "000000000002",
        "UNIVERSITAS TERBUKA 2023.2",
        "FHISIP - Sosiologi",
        "000000000000002",
        1850000,
        "22 Januari 2027 Pukul 11.59 WIB",
    ],
]


def generate_master_data_template() -> bytes:
    """Generate empty Excel template file with standard Master Data Mahasiswa columns and sample data."""
    wb = openpyxl.Workbook()
    ws = cast(Worksheet, wb.active)
    ws.title = "Master_Data_Mahasiswa"
    ws.append(MASTER_TEMPLATE_HEADERS)
    for sample in MASTER_SAMPLE_ROWS:
        ws.append(sample)

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        column_index = col[0].column
        assert isinstance(column_index, int)
        col_letter = get_column_letter(column_index)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@dataclass(frozen=True)
class ImportLayout:
    """Descriptor for recognized workbook sheet structure and headers."""

    kind: str
    data_sheet: str
    issue_sheet: str | None
    headers: dict[str, str]
    default_period: str


def build_billing_period(billing_year: object, semester_type: object) -> dict[str, object]:
    """Validate UI period fields and return one canonical billing-period representation."""
    year_text = clean_excel_text(billing_year)
    semester = clean_excel_text(semester_type).casefold()
    if not re.fullmatch(r"\d{4}", year_text):
        raise ValueError("Tahun tagihan harus terdiri dari 4 digit.")
    year = int(year_text)
    if year < 2000 or year > 2100:
        raise ValueError("Tahun tagihan harus berada antara 2000 dan 2100.")
    if semester not in {"ganjil", "genap"}:
        raise ValueError("Semester tagihan wajib dipilih: ganjil atau genap.")
    semester_number = "1" if semester == "ganjil" else "2"
    semester_label = "Ganjil" if semester == "ganjil" else "Genap"
    return {
        "code": f"{year}.{semester_number}",
        "label": f"{year} {semester_label}",
        "billing_year": year,
        "semester_type": semester,
    }


def _amount_to_int(value: str) -> int | None:
    """Extract and convert numeric amount string from spreadsheet cell into integer."""
    cleaned = re.sub(r"\D+", "", clean_excel_text(value))
    if not cleaned or not cleaned.isdigit():
        return None
    return int(cleaned)


def _normalized_headers(headers: list[str]) -> dict[str, str]:
    """Map lowercased header names to original stripped header strings."""
    return {normalize_text(header).casefold(): normalize_text(header) for header in headers if normalize_text(header)}


def _find_header_key(headers: dict[str, str], aliases: tuple[str, ...]) -> str:
    """Find the first matching column header key amongst supported alias variants."""
    for alias in aliases:
        key = alias.casefold()
        if key in headers:
            return headers[key]
    return ""


def _require_headers(workbook: Path) -> ImportLayout:
    """Inspect workbook sheets and identify valid tabular layouts for import parsing."""
    sheet_names = workbook_sheet_names(workbook)
    if "Data Sinkron" in sheet_names and "Data Belum Lengkap" in sheet_names:
        sync_headers = _normalized_headers(read_sheet_headers(workbook, "Data Sinkron"))
        missing = [header for header in REQUIRED_SYNC_HEADERS if header.casefold() not in sync_headers]
        if missing:
            raise ValueError(
                "Struktur header sheet Data Sinkron tidak sesuai. "
                f"Kolom wajib: {', '.join(REQUIRED_SYNC_HEADERS)}. "
                f"Kolom belum ditemukan: {', '.join(missing)}."
            )

        issue_headers = _normalized_headers(read_sheet_headers(workbook, "Data Belum Lengkap"))
        missing_issue_headers = [header for header in OPTIONAL_ISSUE_HEADERS if header.casefold() not in issue_headers]
        if missing_issue_headers:
            raise ValueError(
                "Struktur header sheet Data Belum Lengkap tidak sesuai. "
                f"Kolom wajib: {', '.join(OPTIONAL_ISSUE_HEADERS)}. "
                f"Kolom belum ditemukan: {', '.join(missing_issue_headers)}."
            )
        return ImportLayout("legacy", "Data Sinkron", "Data Belum Lengkap", sync_headers, DEFAULT_PERIOD)

    for sheet_name in sheet_names:
        headers = _normalized_headers(read_sheet_headers(workbook, sheet_name))
        has_nim = any(a in headers for a in ("nim", "nomor induk mahasiswa"))
        has_name = any(a in headers for a in ("nama", "nama mahasiswa", "nama lengkap"))
        has_briva = any(a in headers for a in ("no rek", "no. rek", "no rekening", "briva", "nomor briva", "va"))
        has_amount = any(a in headers for a in ("jumlah", "nominal", "tagihan", "biaya"))
        if has_nim and has_name and has_briva and has_amount:
            return ImportLayout("current", sheet_name, None, headers, DEFAULT_CURRENT_PERIOD)

    raise ValueError(
        "Struktur Excel tidak dikenali. Gunakan format Data Sinkron lama atau format Master Data "
        "dengan kolom NIM, Nama, No Rek, dan Jumlah."
    )


def _record_value(record: dict[str, str], headers: dict[str, str], aliases: str | tuple[str, ...]) -> str:
    """Extract cell value for the first matched header alias."""
    if isinstance(aliases, str):
        aliases = (aliases,)
    for alias in aliases:
        key = alias.casefold()
        if key in headers:
            return record.get(headers[key], "")
    return ""


def _normalize_briva(value: object) -> str:
    """Validate and clean BRIVA virtual account number format."""
    raw = clean_excel_text(value)
    if not raw or not re.fullmatch(r"[\d\s-]+", raw):
        return ""
    return normalize_nim(raw)


def _read_sync_rows(
    workbook: Path,
    layout: ImportLayout,
    period: str,
) -> tuple[list[dict[str, object]], list[dict[str, object]], list[dict[str, object]], int, list[dict[str, object]]]:
    """Parse and normalize all data rows from the target workbook worksheet."""
    rows: list[dict[str, object]] = []
    errors: list[dict[str, object]] = []
    sample: list[dict[str, object]] = []
    identity_conflict_rows = 0
    skipped_issues: list[dict[str, object]] = []

    for record in read_sheet(workbook, layout.data_sheet):
        raw_nim = clean_excel_text(_record_value(record, layout.headers, ("NIM", "Nomor Induk Mahasiswa")))
        try:
            nim = validate_nim_value(raw_nim)
        except ValueError:
            nim = ""
        raw_name = _record_value(record, layout.headers, ("Nama", "Nama Mahasiswa", "Nama Lengkap"))
        raw_briva = _record_value(
            record, layout.headers, ("No Rek", "No. Rek", "No Rekening", "BRIVA", "Nomor BRIVA", "VA")
        )
        raw_amount = _record_value(record, layout.headers, ("Jumlah", "Nominal", "Tagihan", "Biaya"))
        full_name = normalize_imported_name(raw_name)
        briva = _normalize_briva(raw_briva)
        amount = _amount_to_int(raw_amount)
        row_number = int(record.get("_row_number") or 0)
        if not nim or not full_name or not briva or amount is None:
            message = "Baris dilewati karena NIM, nama, BRIVA, atau nominal tidak valid."
            errors.append(
                {
                    "sheet": layout.data_sheet,
                    "sheet_name": layout.data_sheet,
                    "row_number": row_number,
                    "severity": "warning",
                    "issue_code": "INVALID_REQUIRED_FIELD",
                    "message": message,
                    "note": message,
                    "nim": raw_nim,
                    "full_name": clean_excel_text(raw_name),
                    "briva": clean_excel_text(raw_briva),
                    "amount": clean_excel_text(raw_amount),
                }
            )
            skipped_issues.append(
                {
                    "sheet_name": layout.data_sheet,
                    "row_number": row_number,
                    "severity": "warning",
                    "issue_code": "INVALID_REQUIRED_FIELD",
                    "nim": raw_nim,
                    "full_name": clean_excel_text(raw_name),
                    "briva": clean_excel_text(raw_briva),
                    "amount": clean_excel_text(raw_amount),
                    "note": message,
                }
            )
            continue

        raw_ktp = _record_value(record, layout.headers, ("NO KTP", "No. KTP", "NIK", "KTP", "Nomor KTP"))
        raw_tempat = _record_value(record, layout.headers, ("Tempat Lahir", "Tempat_Lahir"))
        raw_tgl = _record_value(record, layout.headers, ("Tanggal Lahir", "Tanggal_Lahir", "Tgl Lahir"))
        raw_ibu = _record_value(record, layout.headers, ("Nama Ibu Kandung", "Nama Ibu", "Ibu Kandung"))
        raw_email = _record_value(record, layout.headers, ("e-Mail", "Email", "E-mail", "Surel"))
        raw_kontak = _record_value(
            record, layout.headers, ("No Kontak", "No. Kontak", "No Hp", "No. HP", "No Telepon", "Telepon")
        )
        raw_reg = _record_value(record, layout.headers, ("Registrasi Awal", "Periode Masuk", "Registrasi_Awal"))
        raw_prodi = _record_value(record, layout.headers, ("Program Studi", "Prodi", "Jurusan"))
        raw_due = _record_value(record, layout.headers, ("Batas Pembayaran", "Jatuh Tempo", "Due Date"))

        initial_reg = clean_demographic_value(raw_reg)
        entry_year, entry_semester, entry_period = parse_entry_registration(initial_reg)

        row = {
            "nim": nim,
            "full_name": full_name,
            "briva": briva,
            "amount": amount,
            "row_number": row_number,
            "period": period,
            "no_ktp": clean_demographic_value(raw_ktp),
            "tempat_lahir": clean_demographic_value(raw_tempat),
            "tanggal_lahir": clean_demographic_value(raw_tgl),
            "nama_ibu_kandung": clean_demographic_value(raw_ibu),
            "email": clean_demographic_value(raw_email),
            "phone_number": normalize_nim(raw_kontak) if raw_kontak and clean_demographic_value(raw_kontak) else None,
            "initial_registration": initial_reg,
            "entry_year": entry_year,
            "entry_semester": entry_semester,
            "entry_period": entry_period,
            "program_study": clean_demographic_value(raw_prodi),
            "due_date": clean_demographic_value(raw_due),
        }
        rows.append(row)
        if len(sample) < 5:
            sample.append(
                {key: row[key] for key in ("nim", "full_name", "briva", "amount", "program_study", "due_date")}
            )

    canonical_profiles: dict[str, dict[str, object]] = {}
    for row in rows:
        nim = str(row["nim"])
        canonical = canonical_profiles.get(nim)
        if canonical is None:
            canonical_profiles[nim] = row
            continue
        if normalize_name(str(canonical["full_name"])) != normalize_name(str(row["full_name"])):
            identity_conflict_rows += 1
            errors.append(
                {
                    "sheet": layout.data_sheet,
                    "sheet_name": layout.data_sheet,
                    "row_number": row["row_number"],
                    "severity": "warning",
                    "issue_code": "IDENTITY_NAME_MISMATCH",
                    "message": "NIM muncul dengan nama berbeda. Nama pada baris pertama digunakan untuk profil mahasiswa.",
                    "note": "NIM muncul dengan nama berbeda. Nama pada baris pertama digunakan untuk profil mahasiswa.",
                    "nim": row["nim"],
                    "full_name": row["full_name"],
                    "briva": row["briva"],
                    "amount": row["amount"],
                }
            )
        for field in (
            "full_name",
            "no_ktp",
            "tempat_lahir",
            "tanggal_lahir",
            "nama_ibu_kandung",
            "email",
            "phone_number",
            "program_study",
            "initial_registration",
            "entry_year",
            "entry_semester",
            "entry_period",
        ):
            if canonical.get(field):
                row[field] = canonical[field]

    return rows, errors, sample, identity_conflict_rows, skipped_issues
